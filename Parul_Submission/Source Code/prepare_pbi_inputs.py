from pathlib import Path
import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / 'data'
PROCESSED_DIR = DATA_DIR / 'processed'
OUTPUT_DIR = PROJECT_ROOT / 'powerbi_report' / 'data'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

files = {
    'dim_fund': DATA_DIR / '01_fund_master.csv',
    'fact_nav': PROCESSED_DIR / 'nav_history_clean.csv',
    'fact_transactions': PROCESSED_DIR / 'investor_transactions_clean.csv',
    'fact_performance': PROCESSED_DIR / 'scheme_performance_clean.csv',
}

for name, path in files.items():
    df = pd.read_csv(path)
    out_path = OUTPUT_DIR / f'{name}.csv'
    df.to_csv(out_path, index=False)
    print(f'{name}: {len(df)} rows -> {out_path.name}')

# Create Excel workbook for easier Power BI import
wb = Workbook()
# remove the default sheet and create one per table
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
for ws_name in ['dim_fund', 'fact_nav', 'fact_transactions', 'fact_performance']:
    df = pd.read_csv(files[ws_name])
    ws = wb.create_sheet(ws_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

xlsx_path = OUTPUT_DIR / 'bluestock_powerbi_data.xlsx'
wb.save(xlsx_path)
print('excel workbook', xlsx_path)

# Verify SQLite counts
conn = sqlite3.connect(PROJECT_ROOT / 'bluestock_mf.db')
cur = conn.cursor()
for table in ['dim_fund', 'fact_nav', 'fact_transactions', 'fact_performance']:
    count = cur.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]
    print(table, count)
conn.close()
